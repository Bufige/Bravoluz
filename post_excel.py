from openpyxl import Workbook
from get_products_info import data
from __main__ import FILENAME_EXCEL_OUPUT


# here is the module to add our data to excel.


#make a excel table
wb = Workbook()
#set active
ws = wb.active

#setup our current table
ws['A1'] = 'Title'
ws['B1'] = 'Image'
ws['C1'] = 'Model'
ws['D1'] = 'Categories'
ws['E1'] = 'Description'
ws['F1'] = 'Compatibility'

#position to start
pos = 2

#go though our data and add to the table
for product in data:
	str_pos = str(pos)
	ws['A' + str_pos] = product['title']
	ws['B' + str_pos] = product['image']
	ws['C' + str_pos] = product['model']
	ws['D' + str_pos] = product['categories']
	ws['E' + str_pos] = product['description']
	ws['F' + str_pos] = product['compatibility']

	pos += 1


#custom snippet to fix the width from the columns

dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
for col, value in dims.items():
    ws.column_dimensions[col].width = value

#save our table which will generate our excel file.
wb.save(FILENAME_EXCEL_OUPUT)